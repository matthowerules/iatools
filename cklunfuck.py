#!/usr/bin/env python
#based from ckl2poam, originally created by Allyn Stott

VERSION = 1.0

import sys
from xml.dom.minidom import parseString, Node
from openpyxl import Workbook

# show help
if len(sys.argv) < 3:
    print 'cklunfuck version', VERSION
    print 'Usage:', sys.argv[0], 'input.ckl output.xls'
    sys.exit(1)

#open ckl file from stigviewer
file = open(sys.argv[1], 'r')
data = file.read()
file.close()
dom = parseString(data)

#initialize new excel workbook/worksheet
wb = Workbook()
ws1 = wb.get_active_sheet()
ws1.title = 'dump'

c1Values = ['Vuln ID','Status','Severity','Rule ID','CCI Reference','Rule Name','Discussion','Check','Fix','Comments']
for i in range(len(c1Values)):
    ws1.cell(row = 1, column = i+1).value = c1Values[i]

#fill excel workbook with values from ckl file
rowCounter = 2
vulnTag = dom.getElementsByTagName('VULN')
for vuln in vulnTag:
    status = vuln.getElementsByTagName('STATUS')[0].firstChild.data #

    if status == 'NotAFinding' or status == 'Open': #need to parse all findings, not just opens...
		severity = vuln.getElementsByTagName('ATTRIBUTE_DATA')[1].firstChild.data
		if severity == 'high':
			cat = 'high'
		elif severity == 'medium':
			cat = 'medium'
		elif severity == 'low':
			cat = 'low'
		else:
			cat = ''

		if vuln.getElementsByTagName('ATTRIBUTE_DATA')[0].hasChildNodes():
			vulnId = vuln.getElementsByTagName('ATTRIBUTE_DATA')[0].firstChild.data
		else:
			vulnId = ''

		if vuln.getElementsByTagName('ATTRIBUTE_DATA')[3].hasChildNodes():
			ruleId = vuln.getElementsByTagName('ATTRIBUTE_DATA')[3].firstChild.data
		else:
			ruleId = ''

		if vuln.getElementsByTagName('ATTRIBUTE_DATA')[23].hasChildNodes():
			cciRef = vuln.getElementsByTagName('ATTRIBUTE_DATA')[23].firstChild.data
		else:
			cciRef = ''

		if vuln.getElementsByTagName('ATTRIBUTE_DATA')[5].hasChildNodes():
			ruleTitle = vuln.getElementsByTagName('ATTRIBUTE_DATA')[5].firstChild.data
		else:
			ruleTitle = ''

		if vuln.getElementsByTagName('ATTRIBUTE_DATA')[6].hasChildNodes():
			vulnDiscuss = vuln.getElementsByTagName('ATTRIBUTE_DATA')[6].firstChild.data
		else:
			vulnDiscuss = ''

		if vuln.getElementsByTagName('COMMENTS')[0].hasChildNodes():
			comments = vuln.getElementsByTagName('COMMENTS')[0].firstChild.data
		else:
			comments = ''
		if vuln.getElementsByTagName('ATTRIBUTE_DATA')[8].hasChildNodes():
			checkItem = vuln.getElementsByTagName('ATTRIBUTE_DATA')[8].firstChild.data
		else:
			checkItem = ''
		if vuln.getElementsByTagName('ATTRIBUTE_DATA')[9].hasChildNodes():
			fixItem = vuln.getElementsByTagName('ATTRIBUTE_DATA')[9].firstChild.data
		else:
			fixItem = ''

		sheetVals = [vulnId,status,cat,ruleId,cciRef,ruleTitle,vulnDiscuss,checkItem,fixItem,comments]
		for i in range(len(sheetVals)):
			ws1.cell(row = rowCounter, column = i+1).value = sheetVals[i]
		rowCounter = rowCounter + 1

#write all values to xls file
wb.save(sys.argv[2])
