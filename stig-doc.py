#!/usr/bin/env python

VERSION = 1.0

import sys
from xml.dom.minidom import parseString, Node
from openpyxl import Workbook

# show help
if len(sys.argv) < 3:
    print 'stig-doc version', VERSION
    print 'Usage:', sys.argv[0], 'input.ckl output.xlsx'
    sys.exit(1)

#open ckl file from stigviewer
file = open(sys.argv[1], 'r')
data = file.read()
file.close()
dom = parseString(data)

#initialize new excel workbook/worksheet
wb = Workbook()
ws1 = wb.get_active_sheet()
ws1.title = 'dump'

c1Values = ['Vuln ID','Documentable','Status','Severity','Rule Name','Discussion','Check','Fix','Comments']
for i in range(len(c1Values)):
    ws1.cell(row = 1, column = i+1).value = c1Values[i]

#fill excel workbook with values from ckl file
rowCounter = 2
vulnTag = dom.getElementsByTagName('VULN')
for vuln in vulnTag:
    docable = vuln.getElementsByTagName('ATTRIBUTE_DATA')[12].firstChild.data
    severity = vuln.getElementsByTagName('ATTRIBUTE_DATA')[1].firstChild.data
    if docable == 'true':
        status = vuln.getElementsByTagName('STATUS')[0].firstChild.data #
        if vuln.getElementsByTagName('ATTRIBUTE_DATA')[8].hasChildNodes():
			checkItem = vuln.getElementsByTagName('ATTRIBUTE_DATA')[8].firstChild.data
        else:
            checkItem = ''

        if vuln.getElementsByTagName('ATTRIBUTE_DATA')[9].hasChildNodes():
            fixItem = vuln.getElementsByTagName('ATTRIBUTE_DATA')[9].firstChild.data
        else:
            fixItem = ''

        if vuln.getElementsByTagName('ATTRIBUTE_DATA')[0].hasChildNodes():
            vulnId = vuln.getElementsByTagName('ATTRIBUTE_DATA')[0].firstChild.data
        else:
            vulnId = ''

        if vuln.getElementsByTagName('ATTRIBUTE_DATA')[5].hasChildNodes():
            ruleTitle = vuln.getElementsByTagName('ATTRIBUTE_DATA')[5].firstChild.data
        else:
            ruleTitle = ''

        if vuln.getElementsByTagName('ATTRIBUTE_DATA')[6].hasChildNodes():
            vulnDiscuss = vuln.getElementsByTagName('ATTRIBUTE_DATA')[6].firstChild.data
        else:
            vulnDiscuss = ''

        if vuln.getElementsByTagName('COMMENTS')[0].hasChildNodes():
            comments = vuln.getElementsByTagName('COMMENTS')[0].firstChild.data
        else:
            comments = ''

        sheetVals = [vulnId,docable,status,severity,ruleTitle,vulnDiscuss,checkItem,fixItem,comments]
        for i in range(len(sheetVals)):
            ws1.cell(row = rowCounter, column = i+1).value = sheetVals[i]
        rowCounter = rowCounter + 1

#write all values to xlsx file
wb.save(sys.argv[2])
