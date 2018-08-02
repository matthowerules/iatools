#!/usr/bin/env python
#again, from allyn scott... modded for multiple files
#gets nested and ugly quick (ckls are gross)

VERSION = 0.3

import os, sys
from xml.dom.minidom import Node, parseString
from openpyxl import Workbook

# show help
if len(sys.argv) < 3:
    print 'cklufa version', VERSION
    print 'Usage:', sys.argv[0], 'C:\\input\\folder output.xls'
    sys.exit(1)

#init counter
rowCounter = 2

#initialize new excel workbook/worksheet - do it early out of the loop to not overwrite each iteration
wb = Workbook()
ws1 = wb.get_active_sheet()
ws1.title = 'dump'

#c1Values = ['Source','Vuln ID','Status','Severity','Rule ID','CCI Reference','Rule Name','Discussion','Comments']
c1Values = ['Source','Vuln ID','Status','Severity','Rule ID','Rule Name','Discussion','Comments']
for i in range(0, 8):
    ws1.cell(row = 1, column = i+1).value = c1Values[i]

#recursively walk through subdirs for all ckls
for root, subFolders, files in os.walk(sys.argv[1]):
    for file in files:
        if file.endswith(".ckl"):
            #open ckl file from stigviewer
            with open(os.path.join(root, file), 'r') as fin:
                data = fin.read()
                fin.close() #do i need this still?
                #print os.path.join(root, file) #testing paths... delete later
                dom = parseString(data) #do i need THIS still?

                #fill excel workbook with values from ckl file
                vulnTag = dom.getElementsByTagName('VULN')
                for vuln in vulnTag:
                    status = vuln.getElementsByTagName('STATUS')[0].firstChild.data #

                    if status == 'Open': #prev - parsed all findings, nf and opens... status == 'NotAFinding'
                		severity = vuln.getElementsByTagName('ATTRIBUTE_DATA')[1].firstChild.data
                		if severity == 'high':
                			cat = 'high'
                		elif severity == 'medium':
                			cat = 'medium'
                		elif severity == 'low':
                			cat = 'low'
                		else:
                			cat = ''

                		if vuln.getElementsByTagName('ATTRIBUTE_DATA')[0].hasChildNodes():
                			vulnId = vuln.getElementsByTagName('ATTRIBUTE_DATA')[0].firstChild.data
                		else:
                			vulnId = ''

                		if vuln.getElementsByTagName('ATTRIBUTE_DATA')[3].hasChildNodes():
                			ruleId = vuln.getElementsByTagName('ATTRIBUTE_DATA')[3].firstChild.data
                		else:
                			ruleId = ''

#TODO
#fuck you, multi ccis   if vuln.getElementsByTagName('ATTRIBUTE_DATA')[23].hasChildNodes():
#                			cciRef = vuln.getElementsByTagName('ATTRIBUTE_DATA')[23].firstChild.data
#                		else:
#                			cciRef = ''

                		if vuln.getElementsByTagName('ATTRIBUTE_DATA')[5].hasChildNodes():
                			ruleTitle = vuln.getElementsByTagName('ATTRIBUTE_DATA')[5].firstChild.data
                		else:
                			ruleTitle = ''

                		if vuln.getElementsByTagName('ATTRIBUTE_DATA')[6].hasChildNodes():
                			vulnDiscuss = vuln.getElementsByTagName('ATTRIBUTE_DATA')[6].firstChild.data
                		else:
                			vulnDiscuss = ''

                		if vuln.getElementsByTagName('COMMENTS')[0].hasChildNodes():
                			comments = vuln.getElementsByTagName('COMMENTS')[0].firstChild.data
                		else:
                			comments = ''

#TODO-fix CCIs     		sheetVals = [file,vulnId,status,cat,ruleId,cciRef,ruleTitle,vulnDiscuss,comments]
                		sheetVals = [file,vulnId,status,cat,ruleId,ruleTitle,vulnDiscuss,comments]
                		for i in range(0, 8):
                			ws1.cell(row = rowCounter, column = i+1).value = sheetVals[i]
                		rowCounter = rowCounter + 1
                #write all values to xls file
wb.save(sys.argv[2])
